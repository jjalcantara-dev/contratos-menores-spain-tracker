"""
Versión local del scraper — genera un archivo .xlsx sin configurar Google Sheets.
Uso:
    python scraper_xlsx.py                # año actual
    python scraper_xlsx.py 2023           # año concreto
    python scraper_xlsx.py 2012 2025      # rango completo
    python scraper_xlsx.py 2012-2025      # misma sintaxis con guion
"""

from __future__ import annotations

import re
import sys
import logging
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from scraper_core import (
    scrape, URL_PERFIL, FECHA_DESDE, FECHA_HASTA, YEAR,
    parse_años, preparar_estadisticas, normalizar_fila,
    TAB_ESTADISTICAS, TAB_REGISTRO_TOTAL, PREFIJO_CONTRATO, PATRON_TAB_AÑO,
    FilaContrato, DatosEstadisticas,
)

OUTPUT_DIR  = Path("output")
NOMBRE_XLSX = OUTPUT_DIR / "contratos.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(OUTPUT_DIR / f"contratos_{YEAR}.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constantes de presentación
# ---------------------------------------------------------------------------

VERDE_OSCURO = "1F5C2E"
VERDE_CLARO  = "E8F5E9"
EURO_FMT     = '#,##0.00 "€"'

COL_ANCHOS_AÑO   = [(1, 5),  (2, 40), (3, 14), (4, 80), (5, 16)]
COL_ANCHOS_TOTAL = [(1, 8),  (2, 40), (3, 14), (4, 80), (5, 16)]
COL_ANCHOS_EST   = [(1, 40), (2, 16), (3, 16), (4, 20)]

CHART_W_SM = 15   # cm
CHART_H_SM = 10
CHART_W_LG = 20
CHART_H_LG = 14

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

@dataclass
class Estilos:
    fill_verde:  PatternFill
    fill_claro:  PatternFill
    font_blanco: Font
    center:      Alignment
    left:        Alignment

    @classmethod
    def crear(cls) -> Estilos:
        return cls(
            fill_verde  = PatternFill("solid", fgColor=VERDE_OSCURO),
            fill_claro  = PatternFill("solid", fgColor=VERDE_CLARO),
            font_blanco = Font(bold=True, color="FFFFFF", size=11),
            center      = Alignment(horizontal="center", vertical="center", wrap_text=True),
            left        = Alignment(horizontal="left",   vertical="center", wrap_text=True),
        )

# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _tabs_año_xlsx(wb) -> list[tuple[int, object]]:  # análogo a _tabs_año_sheets en scraper.py
    """Devuelve [(year_int, worksheet), ...] ordenados cronológicamente."""
    resultado = []
    for name in wb.sheetnames:
        m = PATRON_TAB_AÑO.match(name)
        if m:
            resultado.append((int(m.group(1)), wb[name]))
    resultado.sort()
    return resultado


def _leer_datos_año(ws_year, year: int) -> list[FilaContrato]:
    """Lee filas de datos de un tab de año xlsx (excluye meta/cabecera/total)."""
    return [
        fila for row in ws_year.iter_rows(min_row=4, values_only=True)
        if (fila := normalizar_fila(row, year)) is not None
    ]


def _open_or_create_wb():
    if NOMBRE_XLSX.exists():
        return load_workbook(NOMBRE_XLSX)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _reordenar_tabs(wb) -> None:
    """Ordena: Estadísticas → Registro Total → años desc (2026, 2025, ...)"""
    names      = wb.sheetnames
    year_names = sorted(
        [n for n in names if PATRON_TAB_AÑO.match(n)],
        reverse=True,
    )
    desired = [t for t in [TAB_ESTADISTICAS, TAB_REGISTRO_TOTAL] if t in names]
    desired.extend(year_names)

    sheet_map  = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [sheet_map[n] for n in desired if n in sheet_map]


def _escribir_tabla(
    ws,
    estilos: Estilos,
    cabecera: list[str],
    filas: list[list],
    euro_col: int | None = None,
) -> tuple[int, int]:
    """
    Escribe cabecera verde + filas con colores alternos.
    Devuelve (fila_cabecera, fila_último_dato).
    """
    ws.append(cabecera)
    fila_cab = ws.max_row
    for col_idx in range(1, len(cabecera) + 1):
        c = ws.cell(fila_cab, col_idx)
        c.fill = estilos.fill_verde
        c.font = estilos.font_blanco
        c.alignment = estilos.center

    for i, fila in enumerate(filas, 1):
        ws.append(fila)
        fila_actual = ws.max_row
        for col_idx in range(1, len(cabecera) + 1):
            cell = ws.cell(fila_actual, col_idx)
            if i % 2 == 0:
                cell.fill = estilos.fill_claro
            cell.alignment = estilos.left
        if euro_col:
            ws.cell(fila_actual, euro_col).number_format = EURO_FMT

    return fila_cab, ws.max_row


def _escribir_fila_total(ws, estilos: Estilos, n_cols: int, total: float) -> None:
    """Añade fila en blanco + fila TOTAL GENERAL con estilo verde."""
    ws.append([])
    fila: list = [""] * n_cols
    fila[1]  = "TOTAL GENERAL"
    fila[-1] = round(total, 2)
    ws.append(fila)
    fila_total = ws.max_row
    for col_idx in range(1, n_cols + 1):
        c = ws.cell(fila_total, col_idx)
        c.fill = estilos.fill_verde
        c.font = estilos.font_blanco

# ---------------------------------------------------------------------------
# Tab de año
# ---------------------------------------------------------------------------

def _escribir_tab_año(
    wb,
    year: str,
    ranking: list,
    total_global: float,
    paginas_con_error: list,
) -> None:
    """Escribe/sobreescribe el tab de un año en el workbook (sin guardar)."""
    tab_nombre = f"{PREFIJO_CONTRATO} {year}"
    if tab_nombre in wb.sheetnames:
        wb.remove(wb[tab_nombre])
    ws = wb.create_sheet(title=tab_nombre)

    estilos = Estilos.crear()
    hoy     = date.today().strftime("%d/%m/%Y")
    aviso   = f"Páginas con error: {paginas_con_error}" if paginas_con_error else ""
    ws.append([f"Última actualización: {hoy}", f"Total adjudicatarios: {len(ranking)}", aviso])
    ws.append([])

    filas = [
        [i, empresa, len(datos["proyectos"]),
         " | ".join(f"{d} ({v:.2f} €)" for d, v in datos["proyectos"]),
         round(datos["total"], 2)]
        for i, (empresa, datos) in enumerate(ranking, 1)
    ]
    fila_cab, _ = _escribir_tabla(
        ws, estilos,
        ["#", "Adjudicatario", "Nº Contratos", "Proyectos", "Total (€)"],
        filas,
    )
    _escribir_fila_total(ws, estilos, 5, total_global)

    for col, ancho in COL_ANCHOS_AÑO:
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = f"A{fila_cab + 1}"


def escribir_xlsx(ranking: list, total_global: float, paginas_con_error: list) -> None:
    """Punto de entrada para año único (usa YEAR del entorno). Guarda el archivo."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = _open_or_create_wb()
    _escribir_tab_año(wb, YEAR, ranking, total_global, paginas_con_error)
    actualizar_registro_total(wb)
    actualizar_estadisticas(wb)
    _reordenar_tabs(wb)
    wb.save(NOMBRE_XLSX)
    log.info(f"Archivo actualizado: {NOMBRE_XLSX}")

# ---------------------------------------------------------------------------
# Tab Registro Total
# ---------------------------------------------------------------------------

def actualizar_registro_total(wb) -> None:
    """Reconstruye 'Registro Total' leyendo todos los tabs de año."""
    year_tabs = _tabs_año_xlsx(wb)
    if not year_tabs:
        return

    if TAB_REGISTRO_TOTAL in wb.sheetnames:
        wb.remove(wb[TAB_REGISTRO_TOTAL])
    ws = wb.create_sheet(TAB_REGISTRO_TOTAL)

    estilos: Estilos     = Estilos.crear()
    todas_filas: list    = []
    for year, ws_year in year_tabs:
        todas_filas.extend(_leer_datos_año(ws_year, year))
    todas_filas.sort(key=lambda r: r["total"], reverse=True)

    hoy  = date.today().strftime("%d/%m/%Y")
    años = ", ".join(str(y) for y, _ in year_tabs)
    ws.append([f"Última actualización: {hoy}", f"Total registros: {len(todas_filas)}", f"Años: {años}"])
    ws.append([])

    filas = [
        [f["año"], f["empresa"], f["num_contratos"], f["proyectos"], round(f["total"], 2)]
        for f in todas_filas
    ]
    fila_cab, fila_fin = _escribir_tabla(
        ws, estilos,
        ["Año", "Adjudicatario", "Nº Contratos", "Proyectos", "Total (€)"],
        filas,
    )
    _escribir_fila_total(ws, estilos, 5, sum(f["total"] for f in todas_filas))

    for col, ancho in COL_ANCHOS_TOTAL:
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes    = f"A{fila_cab + 1}"
    ws.auto_filter.ref = f"A{fila_cab}:E{fila_fin}"

    log.info(f"Tab 'Registro Total' actualizado: {len(todas_filas)} registros de {len(year_tabs)} año(s)")

# ---------------------------------------------------------------------------
# Tab Estadísticas — construcción, formato y gráficas separados
# ---------------------------------------------------------------------------

def _escribir_tablas_estadisticas(
    ws,
    estilos: Estilos,
    datos: DatosEstadisticas,
) -> tuple[int, int, int, int]:
    """
    Escribe las dos tablas de resumen.
    Devuelve (fila_cab_años, fila_fin_años, fila_cab_top, fila_fin_top).
    """
    font_titulo = Font(bold=True, size=12)

    ws.append(["Resumen por año"])
    ws.cell(ws.max_row, 1).font = font_titulo
    filas_años = [
        [y, round(datos.año_resumen[y]["total"], 2),
         datos.año_resumen[y]["contratos"],
         datos.año_resumen[y]["adjudicatarios"]]
        for y in datos.años_ordenados
    ]
    fila_cab_años, fila_fin_años = _escribir_tabla(
        ws, estilos,
        ["Año", "Total (€)", "Nº Contratos", "Nº Adjudicatarios"],
        filas_años, euro_col=2,
    )

    ws.append([])
    ws.append(["Top 10 adjudicatarios (histórico)"])
    ws.cell(ws.max_row, 1).font = font_titulo
    filas_top = [
        [empresa, round(d["total"], 2), d["contratos"],
         ", ".join(str(y) for y in sorted(d["años"]))]
        for empresa, d in datos.top_empresas
    ]
    fila_cab_top, fila_fin_top = _escribir_tabla(
        ws, estilos,
        ["Empresa", "Total (€)", "Nº Contratos", "Años activos"],
        filas_top, euro_col=2,
    )

    for col, ancho in COL_ANCHOS_EST:
        ws.column_dimensions[get_column_letter(col)].width = ancho

    return fila_cab_años, fila_fin_años, fila_cab_top, fila_fin_top


def _añadir_graficas_xlsx(
    ws,
    fila_cab_años: int,
    fila_fin_años: int,
    fila_cab_top:  int,
    fila_fin_top:  int,
) -> None:
    """Inserta las 4 gráficas en la hoja de Estadísticas."""
    cats_años = Reference(ws, min_col=1, min_row=fila_cab_años + 1, max_row=fila_fin_años)

    def _col_chart(title: str, data_col: int, anchor: str) -> None:
        c = BarChart()
        c.type, c.title, c.style = "col", title, 10
        c.width, c.height = CHART_W_SM, CHART_H_SM
        c.add_data(
            Reference(ws, min_col=data_col, min_row=fila_cab_años, max_row=fila_fin_años),
            titles_from_data=True,
        )
        c.set_categories(cats_años)
        ws.add_chart(c, anchor)

    _col_chart("Gasto total por año (€)",                2, "F2")
    _col_chart("Nº de contratos adjudicados por año",    3, "F22")
    _col_chart("Nº de adjudicatarios distintos por año", 4, "V2")

    c4 = BarChart()
    c4.type, c4.title, c4.style = "bar", "Top 10 adjudicatarios por importe total (€)", 10
    c4.width, c4.height = CHART_W_LG, CHART_H_LG
    c4.x_axis.title = "€"
    c4.add_data(
        Reference(ws, min_col=2, min_row=fila_cab_top, max_row=fila_fin_top),
        titles_from_data=True,
    )
    c4.set_categories(Reference(ws, min_col=1, min_row=fila_cab_top + 1, max_row=fila_fin_top))
    ws.add_chart(c4, f"A{fila_fin_top + 3}")


def actualizar_estadisticas(wb) -> None:
    """Reconstruye 'Estadísticas' con tablas resumen y gráficas."""
    year_tabs = _tabs_año_xlsx(wb)
    if not year_tabs:
        return

    todas_filas: list = []
    for year, ws_year in year_tabs:
        todas_filas.extend(_leer_datos_año(ws_year, year))
    datos = preparar_estadisticas(todas_filas)

    if TAB_ESTADISTICAS in wb.sheetnames:
        wb.remove(wb[TAB_ESTADISTICAS])
    ws = wb.create_sheet(TAB_ESTADISTICAS)

    estilos = Estilos.crear()
    fila_cab_años, fila_fin_años, fila_cab_top, fila_fin_top = \
        _escribir_tablas_estadisticas(ws, estilos, datos)
    _añadir_graficas_xlsx(ws, fila_cab_años, fila_fin_años, fila_cab_top, fila_fin_top)

    log.info(f"Tab 'Estadísticas' actualizado: {datos.n_años} año(s), top {datos.n_top} adjudicatarios")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = _open_or_create_wb()

    if sys.argv[1:]:
        años    = parse_años(sys.argv[1:])
        hoy_año = date.today().year

        scrapeado = False
        for year in años:
            log.info(f"{'─' * 50}")
            log.info(f"Procesando año {year}...")
            f_hasta = "" if year >= hoy_año else f"31-12-{year}"
            resultado = scrape(URL_PERFIL, f"01-01-{year}", f_hasta)
            if not resultado.ranking:
                log.warning(f"Año {year}: sin datos publicados, se omite.")
                continue
            _escribir_tab_año(wb, str(year), resultado.ranking, resultado.total_global, resultado.paginas_con_error)
            log.info(f"Año {year}: {len(resultado.ranking)} adjudicatarios — {resultado.total_global:,.2f} €")
            scrapeado = True

        if not scrapeado:
            log.error("No se obtuvieron datos para ningún año solicitado.")
            sys.exit(1)
    else:
        resultado = scrape(URL_PERFIL, FECHA_DESDE, FECHA_HASTA)
        if not resultado.ranking:
            log.error("No se obtuvo ningún dato. Abortando.")
            sys.exit(1)
        _escribir_tab_año(wb, YEAR, resultado.ranking, resultado.total_global, resultado.paginas_con_error)

    actualizar_registro_total(wb)
    actualizar_estadisticas(wb)
    _reordenar_tabs(wb)
    wb.save(NOMBRE_XLSX)
    log.info("Proceso completado.")
