import os
import sys
import pytest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import scraper_xlsx
from openpyxl import load_workbook

RANKING = [
    ("Empresa A", {"total": 5000.0,  "proyectos": [("Obra 1", 3000.0), ("Obra 2", 2000.0)]}),
    ("Empresa B", {"total": 3200.50, "proyectos": [("Servicio X", 3200.50)]}),
    ("Empresa C", {"total": 800.0,   "proyectos": [("Suministro Z", 800.0)]}),
]
TOTAL_GLOBAL = 9000.50


@pytest.fixture(autouse=True)
def limpiar_archivo(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(scraper_xlsx, "YEAR", "2025")
    monkeypatch.setattr(scraper_xlsx, "NOMBRE_XLSX", scraper_xlsx.OUTPUT_DIR / "contratos.xlsx")
    yield


# ---------------------------------------------------------------------------
# Tab del año
# ---------------------------------------------------------------------------

def test_archivo_creado():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    assert Path("output/contratos.xlsx").exists()


def test_titulo_hoja():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    assert "Contratos 2025" in wb.sheetnames


def test_cabecera_correcta():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    cabecera = [ws.cell(3, col).value for col in range(1, 6)]
    assert cabecera == ["#", "Adjudicatario", "Nº Contratos", "Proyectos", "Total (€)"]


def test_numero_filas_datos():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    empresas = [ws.cell(3 + i, 2).value for i in range(1, 4)]
    assert empresas == ["Empresa A", "Empresa B", "Empresa C"]


def test_orden_ranking():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    totales = [ws.cell(3 + i, 5).value for i in range(1, 4)]
    assert totales == sorted(totales, reverse=True)


def test_fila_total():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    ultima_fila = ws.max_row
    assert ws.cell(ultima_fila, 2).value == "TOTAL GENERAL"
    assert ws.cell(ultima_fila, 5).value == round(TOTAL_GLOBAL, 2)


def test_numero_contratos_por_empresa():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    assert ws.cell(4, 3).value == 2
    assert ws.cell(5, 3).value == 1


def test_sin_paginas_error_no_aviso():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    assert not ws.cell(1, 3).value


def test_con_paginas_error_muestra_aviso():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [3, 5])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Contratos 2025"]
    assert "3" in str(ws.cell(1, 3).value)
    assert "5" in str(ws.cell(1, 3).value)


# ---------------------------------------------------------------------------
# Tab Registro Total
# ---------------------------------------------------------------------------

def test_registro_total_existe():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    assert "Registro Total" in wb.sheetnames


def test_registro_total_tiene_datos():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Registro Total"]
    empresas = {ws.cell(3 + i, 2).value for i in range(1, 4)}
    assert empresas == {"Empresa A", "Empresa B", "Empresa C"}


def test_registro_total_columna_año():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Registro Total"]
    assert str(ws.cell(4, 1).value) == "2025"


def test_registro_total_cabecera():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Registro Total"]
    cabecera = [ws.cell(3, col).value for col in range(1, 6)]
    assert cabecera == ["Año", "Adjudicatario", "Nº Contratos", "Proyectos", "Total (€)"]


def test_registro_total_fila_total():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Registro Total"]
    ultima_fila = ws.max_row
    assert ws.cell(ultima_fila, 2).value == "TOTAL GENERAL"
    assert ws.cell(ultima_fila, 5).value == round(TOTAL_GLOBAL, 2)


def test_registro_total_acumula_dos_años(tmp_path, monkeypatch):
    """Al escribir dos años, Registro Total debe tener datos de ambos."""
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(scraper_xlsx, "NOMBRE_XLSX", scraper_xlsx.OUTPUT_DIR / "contratos.xlsx")

    monkeypatch.setattr(scraper_xlsx, "YEAR", "2024")
    scraper_xlsx.escribir_xlsx(
        [("Empresa D", {"total": 1000.0, "proyectos": [("X", 1000.0)]})], 1000.0, []
    )

    monkeypatch.setattr(scraper_xlsx, "YEAR", "2025")
    scraper_xlsx.escribir_xlsx(
        [("Empresa E", {"total": 2000.0, "proyectos": [("Y", 2000.0)]})], 2000.0, []
    )

    wb = load_workbook("output/contratos.xlsx")
    assert "Contratos 2024" in wb.sheetnames
    assert "Contratos 2025" in wb.sheetnames

    ws = wb["Registro Total"]
    empresas = {
        ws.cell(r, 2).value
        for r in range(4, ws.max_row + 1)
        if ws.cell(r, 2).value and ws.cell(r, 2).value != "TOTAL GENERAL"
    }
    assert empresas == {"Empresa D", "Empresa E"}


# ---------------------------------------------------------------------------
# Tab Estadísticas
# ---------------------------------------------------------------------------

def test_estadisticas_existe():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    assert "Estadísticas" in wb.sheetnames


def test_estadisticas_contiene_año():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Estadísticas"]
    todos_valores = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any(str(v) == "2025" for v in todos_valores if v is not None)


def test_estadisticas_contiene_top_empresa():
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    ws = wb["Estadísticas"]
    todos_valores = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert "Empresa A" in todos_valores


# ---------------------------------------------------------------------------
# Orden de tabs
# ---------------------------------------------------------------------------

def test_estadisticas_antes_que_registro_total():
    """Estadísticas debe aparecer antes que Registro Total."""
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])
    wb = load_workbook("output/contratos.xlsx")
    names = wb.sheetnames
    assert names.index("Estadísticas") < names.index("Registro Total")


def test_orden_estadisticas_registro_años(tmp_path, monkeypatch):
    """Orden: Estadísticas → Registro Total → años desc (2025, 2024, ...)"""
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(scraper_xlsx, "NOMBRE_XLSX", scraper_xlsx.OUTPUT_DIR / "contratos.xlsx")

    monkeypatch.setattr(scraper_xlsx, "YEAR", "2024")
    scraper_xlsx.escribir_xlsx(
        [("Empresa D", {"total": 1000.0, "proyectos": [("X", 1000.0)]})], 1000.0, []
    )
    monkeypatch.setattr(scraper_xlsx, "YEAR", "2025")
    scraper_xlsx.escribir_xlsx(RANKING, TOTAL_GLOBAL, [])

    wb = load_workbook("output/contratos.xlsx")
    names = wb.sheetnames
    assert names[0] == "Estadísticas"
    assert names[1] == "Registro Total"
    assert names[2] == "Contratos 2025"
    assert names[3] == "Contratos 2024"


# ---------------------------------------------------------------------------
# parse_años (desde scraper_core)
# ---------------------------------------------------------------------------

def test_parse_años_sin_args(monkeypatch):
    import scraper_core
    monkeypatch.setattr(scraper_core, "YEAR", "2025")
    from scraper_core import parse_años
    assert parse_años([]) == [2025]


def test_parse_años_un_año():
    from scraper_core import parse_años
    assert parse_años(["2023"]) == [2023]


def test_parse_años_rango_guion():
    from scraper_core import parse_años
    assert parse_años(["2021-2023"]) == [2021, 2022, 2023]


def test_parse_años_dos_args():
    from scraper_core import parse_años
    assert parse_años(["2021", "2023"]) == [2021, 2022, 2023]


def test_parse_años_rango_invalido_sale():
    from scraper_core import parse_años
    with pytest.raises(SystemExit):
        parse_años(["2025-2020"])
